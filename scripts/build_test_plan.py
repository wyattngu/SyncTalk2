"""Generate SyncTalk_Test_Plan.xlsx — business processes + test cases per feature.

Run from project root:
    backend/venv/bin/python scripts/build_test_plan.py

Output: SyncTalk_Test_Plan.xlsx (in project root)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter


# ── Styling ──────────────────────────────────────────────────────────────
BRAND   = "4F46E5"   # indigo-600
ACCENT  = "0EA5E9"   # sky-500
SUCCESS = "16A34A"   # green-600
GRAY1   = "F1F5F9"   # slate-100
GRAY2   = "E2E8F0"   # slate-200
DARK    = "0F172A"   # slate-900
WHITE   = "FFFFFF"

THIN = Side(border_style="thin", color="CBD5E1")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1_FONT  = Font(name="Calibri", size=18, bold=True, color=DARK)
H2_FONT  = Font(name="Calibri", size=13, bold=True, color=BRAND)
TH_FONT  = Font(name="Calibri", size=11, bold=True, color=WHITE)
TD_FONT  = Font(name="Calibri", size=11, color=DARK)
TD_BOLD  = Font(name="Calibri", size=11, bold=True, color=DARK)

TH_FILL = PatternFill("solid", fgColor=BRAND)
ACCENT_FILL = PatternFill("solid", fgColor=ACCENT)
ZEBRA   = PatternFill("solid", fgColor=GRAY1)
GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")  # green-100
RED_FILL   = PatternFill("solid", fgColor="FEE2E2")  # red-100
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")  # amber-100

WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


# ── Data ─────────────────────────────────────────────────────────────────

OVERVIEW_ROWS = [
    ["Project",       "SyncTalk — internal communication platform"],
    ["Stack",         "Flask + Flask-SocketIO + SQLAlchemy + PostgreSQL 16  |  Next.js 16 + TypeScript + Tailwind v4 + Zustand  |  Gemini API"],
    ["Backend",       "http://localhost:5001  (run.py)"],
    ["Frontend",      "http://localhost:4000  (next dev)"],
    ["Database",      "Docker container 'synctalk_db' on host port 5434 (mapped to 5432 in container)"],
    ["Demo password", "Password123! — works for all seeded users"],
    ["Demo users",    "alice, bob, charlie, diana, ethan, SyncBot (all @synctalk.dev)"],
    ["How to run",    "./start.sh   →   stop with ./stop.sh"],
    ["How to seed",   "cd backend && ./venv/bin/python seed.py"],
    ["", ""],
    ["Features delivered in this iteration", ""],
    ["Feature 1",     "AI thread summarization + AI-powered semantic search (Gemini structured output)"],
    ["Feature 2",     "Reactions (6 emojis) + @mention notifications"],
    ["Feature 3",     "Markdown + code block rendering for thread + reply content (react-markdown + rehype-sanitize + rehype-highlight)"],
    ["Bonus",         "Functional like + share on thread cards (home feed)"],
]

# ── BUSINESS PROCESSES (per feature) ─────────────────────────────────────

PROCESS_HEADER = [
    "Step", "Actor", "Action", "System Response", "Persistence / Side Effect"
]

PROCESS_F1 = [  # AI summarization + search
    [1, "User", "Opens a thread that has ≥ 2 replies", "Frontend mounts ThreadSummaryCard and calls GET /api/ai/threads/<id>/summary", "—"],
    [2, "Backend", "GET /api/ai/threads/<id>/summary", "If thread.summary cached and summary_reply_count == reply_count → return cached. Else returns null (read endpoint never triggers generation).", "Read of threads.summary, threads.summary_key_points, etc."],
    [3, "User", "Clicks 'Generate summary' button", "POST /api/ai/threads/<id>/summarize", "—"],
    [4, "Backend", "Builds prompt = title + content + replies", "Calls Gemini with response_mime_type=application/json. Parses {summary, key_points[], sentiment}.", "WRITES threads.summary, .summary_key_points (JSON), .summary_sentiment, .summary_reply_count, .summary_generated_at"],
    [5, "User", "Sees TL;DR card with bullet points + sentiment chip", "Card shows 'Auto-generated from N replies'. If reply_count > stored count, 'update available' label appears.", "—"],
    [6, "User", "Clicks 'Refresh'", "POST /api/ai/threads/<id>/summarize?force=true → regenerates", "Overwrites previous summary"],
    [7, "User", "Types in navbar search 'AI chatbot tools' and submits", "Router pushes /?q=AI+chatbot+tools", "—"],
    [8, "Frontend", "useSemanticSearch hook fires", "GET /api/ai/search?q=AI+chatbot+tools", "—"],
    [9, "Backend", "Calls Gemini to extract {keywords[], tag_slugs[], intent}", "Runs SQL: ILIKE on title+content for each keyword, JOIN thread_tags for tag_slugs, ORDER BY keyword-hit-count DESC", "Read-only"],
    [10, "User", "Sees 'AI-interpreted search' card with keyword + tag chips, plus matching threads ranked", "If Gemini fails or no key, falls back to naive keyword split", "—"],
    [11, "User", "Clicks X to clear search", "router.push('/') restores normal feed", "—"],
]

PROCESS_F2 = [  # Reactions + mentions
    [1, "User", "Opens a thread or reply", "useReactions hook calls GET /api/reactions?target_type=...&target_id=...", "Read of reactions table"],
    [2, "User", "Clicks the 'Add reaction' (+) button", "Picker pops out with 6 emojis: 👍 ❤️ 🎉 🤔 😄 🚀", "—"],
    [3, "User", "Picks 👍", "POST /api/reactions {target_type, target_id, emoji=👍}", "—"],
    [4, "Backend", "Reaction service toggles", "If exists → DELETE row (state=removed); else INSERT (state=added). UNIQUE constraint (user_id, target_type, target_id, emoji) prevents duplicates.", "WRITE/DELETE on reactions table"],
    [5, "Backend", "Returns aggregated [{emoji, count, user_ids}] grouped by emoji, sorted by count desc", "—", "—"],
    [6, "User", "Sees emoji pill update — count incremented, pill highlighted in indigo if reacted by viewer", "Picker auto-closes", "—"],
    [7, "User", "Clicks the same emoji pill again", "Toggles off — count decreases, pill returns to neutral style", "DELETE on reactions"],
    [8, "Author A", "Creates a thread or reply with text 'Hey @bob check this out'", "Server-side: thread_service.create_reply → mention_service.notify_mentions", "WRITE on replies table"],
    [9, "Backend", "Regex /@([A-Za-z0-9_]{2,30})/ extracts ['bob']", "Looks up bob via UserRepository.find_by_username_ci. Skips: self, unknown, duplicates.", "—"],
    [10, "Backend", "Inserts Notification(type='mention', user_id=bob.id, message='alice mentioned you in a reply', reference_id=thread_id)", "Existing notification:new socket event broadcasts to bob's room", "WRITE on notifications table"],
    [11, "Bob", "Sees red badge on bell icon (Navbar.unreadCount)", "Toast also fires via socket: 'New notification — alice mentioned you in a reply'", "—"],
    [12, "Bob", "Opens /notifications, clicks the mention", "Marks read (PUT /api/notifications/<id>/read), routes to /threads/<reference_id>", "WRITE: notifications.is_read = true"],
]

PROCESS_F3 = [  # Markdown rendering
    [1, "Author", "On /threads/new, types 'Check `npm install` and **bold** text and ```bash\\nls\\n``` block'", "Textarea hint reads 'Markdown supported — **bold**, _italic_, `code`...'", "—"],
    [2, "Author", "Submits", "POST /api/threads — content stored verbatim (raw markdown)", "WRITE on threads.content"],
    [3, "Reader", "Opens /threads/<id>", "Page renders <MarkdownContent>{thread.content}</MarkdownContent>", "—"],
    [4, "Frontend", "MarkdownContent runs react-markdown pipeline", "remark-gfm → tables, strikethrough, task lists. rehype-sanitize → strips <script>, <iframe>, javascript: URLs. rehype-highlight → adds language-* classes. highlight.js github.css → light syntax theme.", "—"],
    [5, "Reader", "Sees rendered output", "**bold** appears bold; `code` in monospace pill; ```bash``` block shown in card-muted bg with syntax highlighting; tables get header bg; quotes get indigo left-border", "—"],
    [6, "Reader", "Pastes content with <script>alert(1)</script>", "rehype-sanitize strips the script tag — only allowed elements + attributes are rendered. Links get default rel/target safety.", "Defense-in-depth XSS protection"],
    [7, "Reader", "Looks at home feed thread card", "Preview text remains plain (line-clamp-2) — markdown collapses badly in 2-line clamp; intentional design choice", "—"],
]

# Tests have shape: [Test ID, Title, Pre-conditions, Steps, Expected, Type, Priority]
TEST_HEADER = [
    "Test ID", "Title", "Pre-conditions", "Steps", "Expected Result", "Type", "Priority"
]

TESTS_F1 = [
    ["F1-T01", "Generate summary on a thread with 3 replies",
     "Logged in as alice. GEMINI_API_KEY configured. Thread 'Welcome to SyncTalk' exists with ≥ 2 replies.",
     "1) Open the thread\n2) Click 'Generate summary'\n3) Wait < 5s",
     "Card displays a 2-3 sentence summary, 3-5 bullet key points, sentiment chip (positive/neutral/mixed/negative), 'Refresh' button.",
     "Functional", "P0"],
    ["F1-T02", "Cached summary is reused on second visit",
     "F1-T01 has been run. Same thread, no new replies.",
     "1) Refresh /threads/<id>\n2) Inspect network tab",
     "GET /api/ai/threads/<id>/summary returns cached payload instantly. No Gemini call. Card shows immediately.",
     "Functional", "P1"],
    ["F1-T03", "Stale indicator appears after a new reply",
     "F1-T01 done. Stored summary_reply_count = 3.",
     "1) Add a new reply (now 4)\n2) Reload thread page",
     "Header subtitle shows 'update available'. is_stale=true in API response. Old summary still visible. Refresh button regenerates.",
     "Functional", "P1"],
    ["F1-T04", "Force refresh regenerates",
     "F1-T01 done.",
     "1) Click Refresh on summary card\n2) Wait",
     "POST /api/ai/threads/<id>/summarize?force=true. summary_generated_at updates. summary text may change.",
     "Functional", "P1"],
    ["F1-T05", "Missing GEMINI_API_KEY returns clear error on POST",
     "Backend started with GEMINI_API_KEY unset.",
     "1) POST /api/ai/threads/<id>/summarize",
     "503 (or 400) with message 'GEMINI_API_KEY is not configured'. UI shows error string in summary card.",
     "Negative", "P1"],
    ["F1-T06", "Summary card hidden on threads with < 2 replies",
     "Thread with 0 or 1 replies.",
     "1) Open thread page",
     "ThreadSummaryCard renders nothing (component returns null when replyCount<2 AND no summary).",
     "UI", "P2"],
    ["F1-T07", "AI search routes from navbar",
     "Logged in.",
     "1) Type 'AI chatbot tools' in navbar search\n2) Press Enter",
     "Router pushes /?q=AI+chatbot+tools. AI-interpreted search panel appears with chips. Threads list filters to relevant matches.",
     "Functional", "P0"],
    ["F1-T08", "AI search extracts tag slugs correctly",
     "Threads tagged 'AI' and 'Technology' exist.",
     "1) Search 'tell me about chatbot development'",
     "filters.tag_slugs contains ['ai'] and/or ['technology']. Threads in those tags rank above others.",
     "Functional", "P1"],
    ["F1-T09", "Empty search returns plain feed",
     "—",
     "1) Visit /?q=",
     "inAiMode is false (empty trim). Normal Latest Threads card. No AI panel.",
     "Functional", "P2"],
    ["F1-T10", "AI search fallback works without API key",
     "GEMINI_API_KEY unset.",
     "1) Search 'gaming'",
     "filters.used_ai = false. Naive keyword filter still returns Gaming-tagged threads.",
     "Negative", "P1"],
    ["F1-T11", "Clear button removes search state",
     "On /?q=foo with results.",
     "1) Click X in search insight panel",
     "URL becomes /. Insight panel disappears. Normal feed shown.",
     "UI", "P2"],
    ["F1-T12", "Summary HTTP GET endpoint is public (read-only)",
     "—",
     "1) curl GET /api/ai/threads/<id>/summary (no Authorization header)",
     "Returns 200 with cached summary or null. No 401.",
     "Authorization", "P2"],
    ["F1-T13", "Summary POST requires authentication",
     "—",
     "1) curl POST /api/ai/threads/<id>/summarize (no token)",
     "401 Unauthorized.",
     "Authorization", "P0"],
]

TESTS_F2 = [
    ["F2-T01", "Add reaction on a thread",
     "Logged in as alice. On /threads/<id>.",
     "1) Click + button under thread\n2) Pick 👍",
     "Pill '👍 1' appears, highlighted indigo. POST /api/reactions returned {state:added}. Picker closes.",
     "Functional", "P0"],
    ["F2-T02", "Toggle off existing reaction",
     "F2-T01 done.",
     "1) Click the existing 👍 pill",
     "Pill disappears (count was 1) or count drops by 1. State=removed.",
     "Functional", "P0"],
    ["F2-T03", "Multiple users react with same emoji",
     "Alice already reacted 👍.",
     "1) Sign in as bob\n2) React 👍 on the same thread",
     "Pill shows '👍 2'. Both alice and bob in user_ids array. Pill highlighted for bob (it's his reaction too).",
     "Functional", "P0"],
    ["F2-T04", "Different emojis tracked separately",
     "Logged in.",
     "1) React 👍\n2) React ❤️",
     "Two pills shown — '👍 1' and '❤️ 1'. Both highlighted for the viewer.",
     "Functional", "P0"],
    ["F2-T05", "UNIQUE constraint prevents duplicate INSERT race",
     "—",
     "1) Send two simultaneous POST /api/reactions for same emoji",
     "Database UNIQUE(user_id, target_type, target_id, emoji) enforces only one row. Second request flips it off (toggle behavior).",
     "Concurrency", "P2"],
    ["F2-T06", "Disallowed emoji rejected",
     "—",
     "1) POST /api/reactions {emoji:'💀'}",
     "400 with 'Emoji not in allowed set'.",
     "Negative", "P1"],
    ["F2-T07", "Reactions on replies work independently",
     "Thread has reply.",
     "1) React on the reply (not the thread)",
     "Reactions for target_type='reply' shown only under that reply. Thread-level reactions unaffected.",
     "Functional", "P0"],
    ["F2-T08", "Anonymous user cannot react",
     "Not logged in.",
     "1) Click reaction button",
     "Button is rendered but POST returns 401 (token_required). UI suppresses click via isAuthenticated check.",
     "Authorization", "P0"],
    ["F2-T09", "@mention in reply creates notification",
     "Alice + bob exist.",
     "1) Sign in as alice\n2) Reply on a thread: 'Pinging @bob'",
     "POST /api/threads/<id>/replies returns 201. Bob's GET /api/notifications shows one new entry: type='mention', message='alice mentioned you in a reply'.",
     "Functional", "P0"],
    ["F2-T10", "@mention is case-insensitive",
     "User 'bob' exists.",
     "1) Reply with '@BOB' or '@Bob'",
     "Bob still receives notification. find_by_username_ci uses ILIKE.",
     "Functional", "P1"],
    ["F2-T11", "Self-mention does NOT notify",
     "Logged in as alice.",
     "1) Reply with '@alice ...'",
     "Notifications table unchanged. mention_service skips when target.id == by_user_id.",
     "Negative", "P1"],
    ["F2-T12", "Unknown @username silently ignored",
     "User 'ghost' does NOT exist.",
     "1) Reply with '@ghost ...'",
     "Reply succeeds. No notifications created. No 4xx.",
     "Negative", "P1"],
    ["F2-T13", "Multiple mentions deduped",
     "—",
     "1) Reply 'Hey @bob @bob @bob check this'",
     "extract_usernames sets() unique. Only ONE notification for bob.",
     "Functional", "P1"],
    ["F2-T14", "Mention in thread title also notifies",
     "—",
     "1) Create thread title='Help needed' content='@charlie can you...'",
     "Charlie gets a 'mentioned you in a thread' notification with reference_id=thread.id.",
     "Functional", "P1"],
    ["F2-T15", "Bell badge updates in real-time via socket",
     "Bob is online with /notifications open.",
     "1) Alice mentions bob",
     "Bob's navbar badge increments without reload (notification:new socket event).",
     "Realtime", "P1"],
]

TESTS_F3 = [
    ["F3-T01", "Bold + italic render correctly",
     "—",
     "1) Create thread with content '**hello** _world_'\n2) Open detail page",
     "'hello' rendered <strong>; 'world' rendered <em>. Layout uses text-foreground.",
     "Functional", "P0"],
    ["F3-T02", "Inline code rendered with monospace pill",
     "—",
     "1) Content includes 'run `npm install`'",
     "'npm install' rendered in monospace, with bg-secondary border-border padding.",
     "Functional", "P0"],
    ["F3-T03", "Fenced code block with language gets syntax highlighting",
     "—",
     "1) Content '```python\\ndef foo():\\n    pass\\n```'",
     "Code rendered in card-muted background. 'def' highlighted as keyword (highlight.js github theme). Whitespace preserved.",
     "Functional", "P0"],
    ["F3-T04", "GFM tables render",
     "—",
     "1) Content with '| H1 | H2 |\\n|----|----|\\n| a | b |'",
     "Real <table> with bordered cells, header in bg-secondary.",
     "Functional", "P1"],
    ["F3-T05", "Task lists render with checkboxes",
     "—",
     "1) Content '- [ ] todo\\n- [x] done'",
     "Two checkboxes (the [x] one is checked). remark-gfm enables this.",
     "Functional", "P2"],
    ["F3-T06", "<script> tag is stripped (XSS defense)",
     "—",
     "1) Content '<script>alert(1)</script>Hello'",
     "Script tag removed by rehype-sanitize. Only 'Hello' renders. No JS execution.",
     "Security", "P0"],
    ["F3-T07", "javascript: URL in link is removed",
     "—",
     "1) Content '[bad](javascript:alert(1))'",
     "Link rendered without href, or stripped entirely by sanitizer.",
     "Security", "P0"],
    ["F3-T08", "Same content also renders in replies",
     "—",
     "1) Reply with '**bold reply**'",
     "Comment body uses MarkdownContent — bold renders. Same XSS protections apply.",
     "Functional", "P1"],
    ["F3-T09", "Plain content unchanged",
     "—",
     "1) Plain text 'Just a normal paragraph.'",
     "Renders as <p> with text-sm text-foreground my-2. No artifacts.",
     "Regression", "P1"],
    ["F3-T10", "Thread card preview stays plain",
     "Thread content has heavy markdown.",
     "1) View home feed",
     "ThreadCard.preview is <p className='line-clamp-2'>...plain text...</p>. Markdown is NOT rendered (intentional).",
     "UI", "P2"],
    ["F3-T11", "Long inline code does not break layout",
     "—",
     "1) Content with long `aaaaaaaaaaaaaaaaaaaaaaaaaa...`",
     "Code wraps or scrolls inside its container. No horizontal page scroll.",
     "UI", "P2"],
    ["F3-T12", "Markdown hint visible on /threads/new",
     "—",
     "1) Visit /threads/new",
     "Below content textarea: 'Markdown supported — **bold**, _italic_, `code`, ```code blocks```, lists, links, tables.'",
     "UI", "P2"],
    ["F3-T13", "External image markdown renders with border + max-width",
     "—",
     "1) Content '![alt](https://picsum.photos/200)'",
     "Image displayed inside markdown body, rounded-lg border. max-w-full prevents overflow.",
     "Functional", "P2"],
]

TESTS_BONUS = [
    ["B-T01", "Like a thread from the home feed",
     "Logged in. On /.",
     "1) Click heart on a thread card",
     "Heart fills, count increments. Optimistic update. POST /api/threads/<id>/like succeeds.",
     "Functional", "P0"],
    ["B-T02", "Unlike rolls back",
     "B-T01 done.",
     "1) Click the same heart again",
     "Heart unfills, count decrements. DELETE /api/threads/<id>/unlike.",
     "Functional", "P0"],
    ["B-T03", "Optimistic UI rolls back on error",
     "Network blocked.",
     "1) Click heart\n2) Force network failure",
     "UI flips temporarily then reverts. Toast 'Could not update like' shown.",
     "Negative", "P1"],
    ["B-T04", "Anonymous user prompted to sign in",
     "Not logged in.",
     "1) Click heart on a card",
     "Toast 'Please sign in to like posts' — no API call fires.",
     "Authorization", "P1"],
    ["B-T05", "Share button copies URL on desktop",
     "Browser without navigator.share (e.g. desktop Chrome).",
     "1) Click Share",
     "navigator.clipboard.writeText fires. Toast 'Link copied' with the URL.",
     "Functional", "P1"],
    ["B-T06", "Share button uses native share on mobile",
     "Browser supporting navigator.share.",
     "1) Click Share",
     "Native OS share sheet appears with title + content excerpt + URL.",
     "Functional", "P2"],
    ["B-T07", "Click on heart does not navigate to thread",
     "—",
     "1) Click heart icon",
     "Heart toggles. URL stays on /. Click is stopPropagation'd from the parent <Link>.",
     "Regression", "P1"],
]


# ── Helpers ──────────────────────────────────────────────────────────────

def write_overview(ws):
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "SyncTalk — Test Plan & Business Processes"
    ws["A1"].font = Font(name="Calibri", size=22, bold=True, color=BRAND)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "Generated for the 3 features delivered in this iteration."
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="64748B")
    ws.merge_cells("A2:E2")

    row = 4
    for label, value in OVERVIEW_ROWS:
        ws.cell(row=row, column=1, value=label).font = TD_BOLD
        ws.cell(row=row, column=1).fill = ZEBRA
        ws.cell(row=row, column=1).alignment = WRAP
        ws.cell(row=row, column=2, value=value).font = TD_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    ws.column_dimensions["A"].width = 26
    for col in "BCDE":
        ws.column_dimensions[col].width = 25

    # Quick legend block
    row += 2
    ws.cell(row=row, column=1, value="Priority legend").font = H2_FONT
    row += 1
    legend = [
        ("P0", "Critical — must pass before release"),
        ("P1", "High — known issues are blockers"),
        ("P2", "Medium — should pass; can ship with workaround"),
    ]
    for pri, desc in legend:
        ws.cell(row=row, column=1, value=pri).font = TD_BOLD
        ws.cell(row=row, column=2, value=desc).font = TD_FONT
        row += 1


def style_table(ws, header_row, n_cols, n_rows):
    # Header
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = TH_FILL
        cell.font = TH_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER_ALL
    ws.row_dimensions[header_row].height = 28

    # Body — zebra + borders
    for r in range(header_row + 1, header_row + 1 + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = TD_FONT
            cell.alignment = WRAP
            cell.border = BORDER_ALL
            if (r - header_row) % 2 == 0:
                cell.fill = ZEBRA


def add_feature_sheet(wb, *, sheet_name, title, summary,
                      process_rows, test_rows):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = H1_FONT
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = summary
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="475569")
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 36

    # Business process header
    ws["A4"] = "Business process"
    ws["A4"].font = H2_FONT
    ws.merge_cells("A4:G4")

    header_row = 5
    for c, h in enumerate(PROCESS_HEADER, start=1):
        ws.cell(row=header_row, column=c, value=h)
    for i, row in enumerate(process_rows, start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(row=header_row + i, column=c, value=v)
    style_table(ws, header_row, len(PROCESS_HEADER), len(process_rows))

    # Test cases section
    test_header_row = header_row + len(process_rows) + 3
    ws.cell(row=test_header_row - 1, column=1, value="Test cases").font = H2_FONT
    ws.merge_cells(start_row=test_header_row - 1, start_column=1,
                   end_row=test_header_row - 1, end_column=7)

    for c, h in enumerate(TEST_HEADER, start=1):
        ws.cell(row=test_header_row, column=c, value=h)

    for i, row in enumerate(test_rows, start=1):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=test_header_row + i, column=c, value=v)

    style_table(ws, test_header_row, len(TEST_HEADER), len(test_rows))

    # Color the Type + Priority columns
    for i in range(1, len(test_rows) + 1):
        type_cell = ws.cell(row=test_header_row + i, column=6)
        pri_cell  = ws.cell(row=test_header_row + i, column=7)
        # Type tinting
        t = (type_cell.value or "").lower()
        if "security" in t:
            type_cell.fill = RED_FILL
        elif "negative" in t:
            type_cell.fill = YELLOW_FILL
        elif "regression" in t or "ui" in t:
            type_cell.fill = ZEBRA
        else:
            type_cell.fill = GREEN_FILL
        type_cell.alignment = CENTER
        # Priority tinting
        p = (pri_cell.value or "").upper()
        if p == "P0":
            pri_cell.fill = RED_FILL
            pri_cell.font = TD_BOLD
        elif p == "P1":
            pri_cell.fill = YELLOW_FILL
        elif p == "P2":
            pri_cell.fill = GREEN_FILL
        pri_cell.alignment = CENTER

    # Column widths
    widths = {
        "A": 12,  # Step / Test ID
        "B": 14,  # Actor / Title
        "C": 36,  # Action / Pre-conditions
        "D": 40,  # System / Steps
        "E": 40,  # Persistence / Expected
        "F": 14,  # Type
        "G": 10,  # Priority
    }
    # When the row is a process row, columns mean different things. Pick wider for the union.
    final_widths = {"A": 12, "B": 18, "C": 38, "D": 42, "E": 42, "F": 14, "G": 10}
    for col_letter, w in final_widths.items():
        ws.column_dimensions[col_letter].width = w


def add_summary_sheet(wb):
    ws = wb.create_sheet("Test Coverage Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Test Coverage Summary"
    ws["A1"].font = H1_FONT
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    headers = ["Feature", "Total", "P0", "P1", "P2", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)

    rows = []
    for label, ts in [
        ("F1 — AI summary + search", TESTS_F1),
        ("F2 — Reactions + mentions", TESTS_F2),
        ("F3 — Markdown rendering",   TESTS_F3),
        ("Bonus — Like + share",      TESTS_BONUS),
    ]:
        p0 = sum(1 for t in ts if t[6] == "P0")
        p1 = sum(1 for t in ts if t[6] == "P1")
        p2 = sum(1 for t in ts if t[6] == "P2")
        rows.append([label, len(ts), p0, p1, p2,
                     f"{p0} blocking · {p1} high · {p2} medium"])

    totals = ["TOTAL",
              sum(r[1] for r in rows),
              sum(r[2] for r in rows),
              sum(r[3] for r in rows),
              sum(r[4] for r in rows),
              ""]

    for i, row in enumerate(rows + [totals], start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(row=3 + i, column=c, value=v)

    style_table(ws, 3, len(headers), len(rows) + 1)

    # Bold totals row
    last = 3 + len(rows) + 1
    for c in range(1, len(headers) + 1):
        ws.cell(row=last, column=c).font = TD_BOLD
        ws.cell(row=last, column=c).fill = ACCENT_FILL
        ws.cell(row=last, column=c).font = Font(name="Calibri", size=11,
                                                bold=True, color=WHITE)

    widths = {"A": 32, "B": 8, "C": 8, "D": 8, "E": 8, "F": 36}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def main():
    wb = Workbook()
    write_overview(wb.active)

    add_feature_sheet(
        wb,
        sheet_name="F1 — AI Summary & Search",
        title="Feature 1 — AI Thread Summarization + Semantic Search",
        summary=("Gemini 2.5 Flash with structured JSON output produces a 2-3 sentence summary, "
                 "3-5 key points, and a sentiment label per thread. Result is cached on threads.summary "
                 "and invalidated when reply_count changes. Free-text search is interpreted by Gemini "
                 "into {keywords, tag_slugs, intent} and ranked via SQL ILIKE + tag joins."),
        process_rows=PROCESS_F1,
        test_rows=TESTS_F1,
    )

    add_feature_sheet(
        wb,
        sheet_name="F2 — Reactions & Mentions",
        title="Feature 2 — Reactions + @Mentions",
        summary=("6-emoji reaction picker per thread + per reply, stored in a UNIQUE-constrained "
                 "reactions table with optimistic toggle UI. @mentions in thread/reply content are "
                 "regex-extracted and resolved case-insensitively, generating type='mention' "
                 "notifications that flow through the existing socket pipeline."),
        process_rows=PROCESS_F2,
        test_rows=TESTS_F2,
    )

    add_feature_sheet(
        wb,
        sheet_name="F3 — Markdown rendering",
        title="Feature 3 — Markdown + Code-block rendering",
        summary=("Thread bodies and reply bodies render Markdown via react-markdown + remark-gfm + "
                 "rehype-sanitize + rehype-highlight (highlight.js github theme). All HTML is sanitized "
                 "with an allowlist; code blocks get syntax highlighting; tables, task lists, and "
                 "strikethrough are supported. Card previews remain plain text by design."),
        process_rows=PROCESS_F3,
        test_rows=TESTS_F3,
    )

    add_feature_sheet(
        wb,
        sheet_name="Bonus — Like & Share",
        title="Bonus — Functional Like + Share on Thread Cards",
        summary=("Heart toggles on the home feed with optimistic update + rollback on error. "
                 "Share button uses navigator.share when available, otherwise falls back to "
                 "navigator.clipboard.writeText with a toast confirmation. Click handlers stopPropagation "
                 "so the parent <Link> does not navigate when the user is just liking/sharing."),
        process_rows=[
            [1, "User", "Hovers a thread card on /", "Card highlights; heart, comment, share row visible at bottom", "—"],
            [2, "User", "Clicks heart icon", "stopPropagation prevents <Link> navigation. Optimistic flip in state.", "—"],
            [3, "Frontend", "Calls threadsService.like or .unlike depending on prior state", "POST /api/threads/<id>/like  or  DELETE /api/threads/<id>/unlike", "WRITE on likes table"],
            [4, "Backend", "Returns {like_count, liked}", "Frontend syncs count from server response", "—"],
            [5, "User", "Clicks Share", "navigator.share if available else navigator.clipboard.writeText. Toast 'Link copied'.", "Clipboard write"],
            [6, "User", "Clicks comment count", "Routes to /threads/<id> via existing <Link>", "Marks thread as viewed in localStorage"],
        ],
        test_rows=TESTS_BONUS,
    )

    add_summary_sheet(wb)

    out = Path(__file__).resolve().parents[1] / "SyncTalk_Test_Plan.xlsx"
    wb.save(out)
    print(f"✓ Wrote {out}")


if __name__ == "__main__":
    main()
